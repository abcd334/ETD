import openpyxl
from requests_html import HTMLSession
from bs4 import BeautifulSoup
import requests
import pandas as pd
import json
from selenium import webdriver

#讀取excel
wb = openpyxl.load_workbook('ETD list.xlsx')
ws = wb['Sheet1']
k=0
for i in range(118,167):  #ws.max_row+1):
    #讀取股票代號
    stock_code = ws.cell(i, 1).value
    url_1 = 'https://www.moneydj.com/us/preferred/pf0003/' + stock_code

    history = pd.read_html(url_1)[1]
    #print(url_1)
    #print(len(history))
    #break
    if len(history)>3:
        #
        #print(history['除息日'][0])
        #


        chrome = webdriver.Chrome('./chromedriver')
        chrome.get(url_1)
        soup = BeautifulSoup(chrome.page_source, "lxml")
        DVD = soup.find_all('span')
        # print(dfs_2)
        DVD_1 = []
        price_1 = ""



        for j in DVD:
            DVD_1.append(j.text)
        # print(len(price))
        #print(DVD_1)
        for j in range(50,80):

            if DVD_1[j] == "全年配息:":
                price_1 = DVD_1[j + 1]

                ws.cell(i, 19, price_1)
                ws.cell(i, 20, history['除息日'][0])

                print(i, stock_code, price_1)
                wb.save('ETD list.xlsx')