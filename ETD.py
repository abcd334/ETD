#爬每年股利
import openpyxl
from requests_html import HTMLSession
from bs4 import BeautifulSoup
import requests
import pandas as pd
import json
from selenium import webdriver

#讀取excel
wb = openpyxl.load_workbook('ETD list.xlsx')
ws = wb['Sheet1']


for i in range(137,186):  #ws.max_row+1):
    #讀取股票代號
    stock_code = str(ws.cell(i, 1).value)[:len(ws.cell(i, 1).value)-1].replace('-', '')
    url_1 = 'https://www.moneydj.com/us/basic/basic0001/' + stock_code
    url_2 = 'https://www.moneydj.com/us/basic/basic0002/' + stock_code

    dfs_1 = pd.read_html(url_1,encoding='utf-8')[1]
    if len(dfs_1)==8:

        chrome = webdriver.Chrome('./chromedriver')
        chrome.get(url_2)
        soup = BeautifulSoup(chrome.page_source, "lxml")
        dfs_2 = soup.find_all('span')
        #print(dfs_2)
        price = []
        price_1=""
        for j in dfs_2:
            price.append(j.text)
        #print(len(price))
        for j in range(len(price)):

            if price[j][0:2]=="成交":
                price_1=price[j+1]

        #price_str=float(str(price[1])[32:38])

        ws.cell(i, 14, stock_code)
        ws.cell(i, 15, dfs_1[3][5])
        ws.cell(i, 16, price_1)
        print(i,stock_code,dfs_1[3][5],price_1)
        wb.save('ETD list.xlsx')

